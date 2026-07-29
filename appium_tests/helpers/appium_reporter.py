"""
MedCare+ Appium Excel Reporter (Python)
Generates timestamped .xlsx reports + JSON sidecar for the HTML report.
"""
import os
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config.appium_config import REPORTS_DIR, PLATFORM


# ── colour constants ──────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1E40AF")   # blue
_PASS_FILL    = PatternFill("solid", fgColor="16A34A")   # green
_FAIL_FILL    = PatternFill("solid", fgColor="DC2626")   # red
_WHITE_FONT   = Font(color="FFFFFF", bold=True)
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=12)


def _style_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


class AppiumExcelReporter:
    """Generates an Excel workbook with Executive Summary + Detailed Results,
    and a JSON sidecar consumed by the consolidated HTML report."""

    def generate_report(
        self,
        results: list[dict],
        summary: dict,
        platform: str = PLATFORM,
    ) -> dict[str, str]:
        """
        Build and save the Excel report + JSON sidecar.

        Parameters
        ----------
        results  : list of result dicts (suiteName, title, status, durationMs, error, timestamp)
        summary  : dict with keys total, passed, failed, durationMs
        platform : 'android' | 'ios'

        Returns
        -------
        dict with keys 'report_path', 'latest_report_path', and 'json_path'.
        """
        wb = openpyxl.Workbook()

        # ── Sheet 1: Executive Summary ────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 25

        ws_summary.append(["Metric", "Value"])
        _style_header(ws_summary)

        total    = summary.get("total", 0)
        passed   = summary.get("passed", 0)
        failed   = summary.get("failed", 0)
        duration = summary.get("durationMs", 0)
        pass_rate = f"{(passed / total * 100):.1f}%" if total > 0 else "0%"

        summary_rows = [
            ("Test Suite",          "MedCare+ Appium Mobile Tests"),
            ("Platform",            platform.upper()),
            ("Execution Date",      datetime.utcnow().isoformat()),
            ("Total Test Cases",    total),
            ("Passed",              passed),
            ("Failed",              failed),
            ("Pass Rate (%)",       pass_rate),
            ("Total Duration (s)",  f"{duration / 1000:.2f}"),
        ]
        for row in summary_rows:
            ws_summary.append(row)

        # ── Sheet 2: Detailed Results ─────────────────────────────────────────
        ws_detail = wb.create_sheet("Test Results")
        col_widths = [6, 28, 42, 10, 16, 50, 24]
        headers    = ["#", "Test Suite", "Test Case", "Status",
                      "Duration (ms)", "Error / Notes", "Timestamp"]

        for col, width in zip("ABCDEFG", col_widths):
            ws_detail.column_dimensions[col].width = width

        ws_detail.append(headers)
        _style_header(ws_detail)

        for idx, r in enumerate(results, start=1):
            status = r.get("status", "UNKNOWN")
            ws_detail.append([
                idx,
                r.get("suiteName", ""),
                r.get("title", ""),
                status,
                r.get("durationMs", 0),
                r.get("error", ""),
                r.get("timestamp", datetime.utcnow().isoformat()),
            ])
            status_cell = ws_detail.cell(row=idx + 1, column=4)
            status_cell.fill = _PASS_FILL if status == "PASS" else _FAIL_FILL
            status_cell.font = _WHITE_FONT

        # ── Save Excel files ──────────────────────────────────────────────────
        os.makedirs(REPORTS_DIR, exist_ok=True)
        ts          = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S")
        report_path = os.path.join(REPORTS_DIR, f"Appium_{platform}_Report_{ts}.xlsx")
        latest_path = os.path.join(REPORTS_DIR, f"Appium_Latest_{platform}.xlsx")

        wb.save(report_path)
        wb.save(latest_path)

        # ── Save JSON sidecar (used by the consolidated HTML report) ─────────
        # Stamp each result with the current platform so the HTML can group them.
        stamped = [{**r, "platform": platform} for r in results]

        # Merge with any existing sidecar from a previous platform run so that
        # when 'both' platforms run, a single JSON file covers all results.
        json_path = os.path.join(REPORTS_DIR, "appium_report_data.json")
        existing: list[dict] = []
        if os.path.exists(json_path):
            try:
                with open(json_path, encoding="utf-8") as fh:
                    existing = json.load(fh)
            except Exception:  # noqa: BLE001
                existing = []

        merged = existing + stamped
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump(merged, fh, indent=2, ensure_ascii=False)

        return {
            "report_path":        report_path,
            "latest_report_path": latest_path,
            "json_path":          json_path,
        }
